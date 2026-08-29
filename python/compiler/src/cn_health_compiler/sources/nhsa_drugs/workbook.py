"""NHSA drug workbook configuration and structural inspection."""

from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Literal, Self
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, model_validator

from cn_health_compiler.core.dataset import load_yaml_mapping
from cn_health_compiler.core.source import SourceSnapshot


class WorkbookContractError(ValueError):
    """Raised when a workbook does not match its declared fingerprint."""


class _ConfigModel(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)


class WorkbookSourceConfig(_ConfigModel):
    filename: str
    sha256: str
    size_bytes: int


class WorkbookStructureConfig(_ConfigModel):
    required_sheets: tuple[str, ...]
    canonical_sheet: str
    resolve_external_links: Literal[False]


class WorkbookContainerConfig(_ConfigModel):
    expected_zip_entries: int
    expected_uncompressed_size_bytes: int
    max_uncompressed_size_bytes: int
    reject_macros: bool

    @model_validator(mode="after")
    def maximum_covers_expected_size(self) -> Self:
        if self.max_uncompressed_size_bytes < self.expected_uncompressed_size_bytes:
            raise ValueError("maximum uncompressed size is below the expected size")
        return self


class WorkbookSheetConfig(_ConfigModel):
    dimension: str
    header_row: int
    first_data_row: int
    expected_data_rows: int
    expected_formula_cells: int
    headers: tuple[str, ...]


class NhsaDrugWorkbookConfig(_ConfigModel):
    version: int
    source: WorkbookSourceConfig
    workbook: WorkbookStructureConfig
    container: WorkbookContainerConfig
    sheet: WorkbookSheetConfig

    @classmethod
    def load(cls, path: Path) -> Self:
        return cls.model_validate(load_yaml_mapping(path))


@dataclass(frozen=True, slots=True)
class WorkbookInspection:
    snapshot: SourceSnapshot
    sheet_names: tuple[str, ...]
    dimension: str
    headers: tuple[str, ...]
    data_rows: int
    formula_cells: int


def inspect_workbook(
    snapshot: SourceSnapshot,
    config: NhsaDrugWorkbookConfig,
) -> WorkbookInspection:
    """Inspect a verified snapshot and fail closed on structural drift."""
    _require_equal("source SHA256", snapshot.sha256, config.source.sha256)
    _require_equal("source size", snapshot.size_bytes, config.source.size_bytes)
    _inspect_container(snapshot.path, config.container)

    workbook = load_workbook(
        snapshot.path,
        read_only=True,
        data_only=False,
        keep_links=config.workbook.resolve_external_links,
    )
    try:
        sheet_names = tuple(workbook.sheetnames)
        _require_equal("sheet names", sheet_names, config.workbook.required_sheets)
        sheet = workbook[config.workbook.canonical_sheet]
        dimension = sheet.calculate_dimension()
        _require_equal("dimension", dimension, config.sheet.dimension)

        header_cells = next(
            sheet.iter_rows(
                min_row=config.sheet.header_row,
                max_row=config.sheet.header_row,
                max_col=len(config.sheet.headers),
            )
        )
        headers = tuple(cell.value for cell in header_cells)
        _require_equal("headers", headers, config.sheet.headers)

        formula_cells = sum(cell.data_type == "f" for cell in header_cells)
        data_rows = 0
        for row in sheet.iter_rows(min_row=config.sheet.first_data_row):
            formula_cells += sum(cell.data_type == "f" for cell in row)
            if any(cell.value not in (None, "") for cell in row):
                data_rows += 1

        _require_equal("data rows", data_rows, config.sheet.expected_data_rows)
        _require_equal("formula cells", formula_cells, config.sheet.expected_formula_cells)
        return WorkbookInspection(
            snapshot=snapshot,
            sheet_names=sheet_names,
            dimension=dimension,
            headers=config.sheet.headers,
            data_rows=data_rows,
            formula_cells=formula_cells,
        )
    finally:
        workbook.close()


def _inspect_container(path: Path, config: WorkbookContainerConfig) -> None:
    try:
        with ZipFile(path) as archive:
            entries = archive.infolist()
            uncompressed_size = sum(entry.file_size for entry in entries)
            if uncompressed_size > config.max_uncompressed_size_bytes:
                raise WorkbookContractError(
                    "workbook exceeds the maximum declared uncompressed size"
                )
            _require_equal("ZIP entries", len(entries), config.expected_zip_entries)
            _require_equal(
                "uncompressed size",
                uncompressed_size,
                config.expected_uncompressed_size_bytes,
            )
            for entry in entries:
                member_path = PurePosixPath(entry.filename)
                if member_path.is_absolute() or ".." in member_path.parts:
                    raise WorkbookContractError(f"unsafe ZIP member path: {entry.filename}")
                if entry.flag_bits & 0x1:
                    raise WorkbookContractError(f"encrypted ZIP member: {entry.filename}")
                if config.reject_macros and entry.filename.casefold().endswith("vbaproject.bin"):
                    raise WorkbookContractError("macro-enabled workbook is not allowed")
            corrupt_member = archive.testzip()
            if corrupt_member is not None:
                raise WorkbookContractError(f"corrupt ZIP member: {corrupt_member}")
    except BadZipFile as error:
        raise WorkbookContractError("source is not a valid XLSX container") from error


def _require_equal(label: str, actual: object, expected: object) -> None:
    if actual != expected:
        raise WorkbookContractError(f"{label} changed: expected {expected!r}, found {actual!r}")
