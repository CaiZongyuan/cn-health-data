"""Raw extraction and canonical normalization for NHSA drug records."""

import unicodedata
from collections.abc import Iterator, Sequence
from dataclasses import dataclass

from openpyxl import load_workbook

from cn_health_compiler.sources.nhsa_drugs.workbook import (
    NhsaDrugWorkbookConfig,
    WorkbookContractError,
    WorkbookInspection,
)

SOURCE_HEADERS = (
    "药品代码",
    "数据来源",
    "注册名称",
    "商品名称",
    "注册剂型",
    "剂型",
    "注册规格",
    "规格",
    "包装材质",
    "最小包装数量",
    "最小制剂单位",
    "最小包装单位",
    "药品企业",
    "分包装企业名称",
    "生产企业",
    "批准文号",
    "原批准文号",
    "药品本位码",
    "上市药品持有人",
    "市场状态",
    "医保药品名称",
    "2025版甲乙类",
    "医保剂型",
    "编号",
    "备注",
    "曾用码",
)

type RawCell = object


class RecordNormalizationError(ValueError):
    """Raised when a source cell cannot be normalized without guessing."""


@dataclass(frozen=True, slots=True)
class RawDrugRow:
    """One source row before normalization."""

    source_row: int
    code: RawCell
    data_source: RawCell
    registered_name: RawCell
    trade_name: RawCell
    registered_dosage_form: RawCell
    dosage_form: RawCell
    registered_specification: RawCell
    specification: RawCell
    packaging_material: RawCell
    minimum_package_quantity: RawCell
    minimum_dosage_unit: RawCell
    minimum_package_unit: RawCell
    drug_company: RawCell
    repackaging_company: RawCell
    manufacturer: RawCell
    approval_number: RawCell
    previous_approval_number: RawCell
    standard_drug_code: RawCell
    marketing_authorization_holder: RawCell
    market_status: RawCell
    insurance_name: RawCell
    reimbursement_class_2025: RawCell
    insurance_dosage_form: RawCell
    insurance_number: RawCell
    note: RawCell
    former_code: RawCell

    @classmethod
    def from_values(cls, source_row: int, values: Sequence[RawCell]) -> "RawDrugRow":
        if len(values) != len(SOURCE_HEADERS):
            raise WorkbookContractError(
                f"source row {source_row} has {len(values)} cells; expected {len(SOURCE_HEADERS)}"
            )
        return cls(
            source_row=source_row,
            code=values[0],
            data_source=values[1],
            registered_name=values[2],
            trade_name=values[3],
            registered_dosage_form=values[4],
            dosage_form=values[5],
            registered_specification=values[6],
            specification=values[7],
            packaging_material=values[8],
            minimum_package_quantity=values[9],
            minimum_dosage_unit=values[10],
            minimum_package_unit=values[11],
            drug_company=values[12],
            repackaging_company=values[13],
            manufacturer=values[14],
            approval_number=values[15],
            previous_approval_number=values[16],
            standard_drug_code=values[17],
            marketing_authorization_holder=values[18],
            market_status=values[19],
            insurance_name=values[20],
            reimbursement_class_2025=values[21],
            insurance_dosage_form=values[22],
            insurance_number=values[23],
            note=values[24],
            former_code=values[25],
        )


@dataclass(frozen=True, slots=True)
class DrugRecord:
    """Canonical drug record ready for validation and persistence."""

    code: str
    data_source: str
    registered_name: str
    trade_name: str
    registered_dosage_form: str
    dosage_form: str
    registered_specification: str
    specification: str
    packaging_material: str
    minimum_package_quantity: str
    minimum_dosage_unit: str
    minimum_package_unit: str
    drug_company: str
    repackaging_company: str | None
    manufacturer: str
    approval_number: str
    previous_approval_number: str | None
    standard_drug_code: str
    marketing_authorization_holder: str | None
    market_status: str
    insurance_name: str | None
    reimbursement_class_2025: str | None
    insurance_dosage_form: str | None
    insurance_number: str | None
    note: str | None
    former_code: str | None
    source_row: int
    source_version: str
    source_sha256: str


DRUG_RECORD_FIELD_NAMES = frozenset(DrugRecord.__dataclass_fields__)


def iter_raw_drug_rows(
    inspection: WorkbookInspection,
    config: NhsaDrugWorkbookConfig,
) -> Iterator[RawDrugRow]:
    """Stream source rows from an already inspected canonical sheet."""
    if config.sheet.headers != SOURCE_HEADERS or inspection.headers != SOURCE_HEADERS:
        raise WorkbookContractError("workbook headers do not match the adapter field mapping")

    workbook = load_workbook(
        inspection.snapshot.path,
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    extracted_rows = 0
    try:
        sheet = workbook[config.workbook.canonical_sheet]
        for source_row, values in enumerate(
            sheet.iter_rows(
                min_row=config.sheet.first_data_row,
                max_col=len(SOURCE_HEADERS),
                values_only=True,
            ),
            start=config.sheet.first_data_row,
        ):
            if not any(value not in (None, "") for value in values):
                continue
            extracted_rows += 1
            yield RawDrugRow.from_values(source_row, values)
        if extracted_rows != inspection.data_rows:
            raise WorkbookContractError(
                f"extracted {extracted_rows} rows after inspection reported {inspection.data_rows}"
            )
    finally:
        workbook.close()


def normalize_raw_drug_row(
    raw: RawDrugRow,
    source_version: str,
    source_sha256: str,
) -> DrugRecord:
    """Normalize one raw row without inferring or correcting source facts."""
    return DrugRecord(
        code=_required_code(raw.code),
        data_source=_required_text(raw.data_source, "data_source"),
        registered_name=_required_text(raw.registered_name, "registered_name"),
        trade_name=_required_text(raw.trade_name, "trade_name"),
        registered_dosage_form=_required_text(raw.registered_dosage_form, "registered_dosage_form"),
        dosage_form=_required_text(raw.dosage_form, "dosage_form"),
        registered_specification=_required_text(
            raw.registered_specification, "registered_specification"
        ),
        specification=_required_text(raw.specification, "specification"),
        packaging_material=_required_text(raw.packaging_material, "packaging_material"),
        minimum_package_quantity=_required_text(
            raw.minimum_package_quantity, "minimum_package_quantity"
        ),
        minimum_dosage_unit=_required_text(raw.minimum_dosage_unit, "minimum_dosage_unit"),
        minimum_package_unit=_required_text(raw.minimum_package_unit, "minimum_package_unit"),
        drug_company=_required_text(raw.drug_company, "drug_company"),
        repackaging_company=_optional_text(raw.repackaging_company, "repackaging_company"),
        manufacturer=_required_text(raw.manufacturer, "manufacturer"),
        approval_number=_required_text(raw.approval_number, "approval_number"),
        previous_approval_number=_optional_text(
            raw.previous_approval_number, "previous_approval_number"
        ),
        standard_drug_code=_required_text(raw.standard_drug_code, "standard_drug_code"),
        marketing_authorization_holder=_optional_text(
            raw.marketing_authorization_holder, "marketing_authorization_holder"
        ),
        market_status=_required_text(raw.market_status, "market_status"),
        insurance_name=_optional_text(raw.insurance_name, "insurance_name"),
        reimbursement_class_2025=_optional_text(
            raw.reimbursement_class_2025, "reimbursement_class_2025"
        ),
        insurance_dosage_form=_optional_text(raw.insurance_dosage_form, "insurance_dosage_form"),
        insurance_number=_optional_text(raw.insurance_number, "insurance_number"),
        note=_optional_text(raw.note, "note"),
        former_code=_optional_text(raw.former_code, "former_code"),
        source_row=raw.source_row,
        source_version=_required_text(source_version, "source_version"),
        source_sha256=_required_text(source_sha256, "source_sha256"),
    )


def _required_code(value: RawCell) -> str:
    if not isinstance(value, str):
        raise RecordNormalizationError("code is required and must be text")
    code = value.strip(" \t\r\n")
    if not code:
        raise RecordNormalizationError("code is required")
    return code


def _required_text(value: RawCell, field_name: str) -> str:
    normalized = _optional_text(value, field_name)
    if normalized is None:
        raise RecordNormalizationError(f"{field_name} is required")
    return normalized


def _optional_text(value: RawCell, field_name: str) -> str | None:
    if value is None:
        return None
    if not isinstance(value, str):
        raise RecordNormalizationError(f"{field_name} must be text, found {type(value).__name__}")
    normalized = unicodedata.normalize("NFC", value.strip())
    return normalized or None
