"""Raw extraction and normalization for NHC clinical diagnosis records."""

import unicodedata
from collections.abc import Iterator, Sequence
from dataclasses import dataclass

from openpyxl import load_workbook

from cn_health_compiler.core.workbook import (
    WorkbookConfig,
    WorkbookContractError,
    WorkbookInspection,
)

SOURCE_HEADERS = ("主要编码", "附加编码", "疾病名称")
type RawCell = object


class DiagnosisNormalizationError(ValueError):
    """Raised when a diagnosis source row cannot be normalized safely."""


@dataclass(frozen=True, slots=True)
class RawDiagnosisRow:
    source_row: int
    main_code: RawCell
    additional_code: RawCell
    name: RawCell

    @classmethod
    def from_values(cls, source_row: int, values: Sequence[RawCell]) -> "RawDiagnosisRow":
        if len(values) != len(SOURCE_HEADERS):
            raise WorkbookContractError(
                f"source row {source_row} has {len(values)} cells; expected {len(SOURCE_HEADERS)}"
            )
        return cls(source_row, values[0], values[1], values[2])


@dataclass(frozen=True, slots=True)
class DiagnosisRecord:
    code: str
    main_code: str | None
    additional_code: str | None
    name: str
    source_row: int
    source_version: str
    source_sha256: str


def iter_raw_diagnosis_rows(
    inspection: WorkbookInspection,
    config: WorkbookConfig,
) -> Iterator[RawDiagnosisRow]:
    if config.sheet.headers != SOURCE_HEADERS or inspection.headers != SOURCE_HEADERS:
        raise WorkbookContractError("workbook headers do not match the diagnosis field mapping")
    workbook = load_workbook(
        inspection.snapshot.path, read_only=True, data_only=False, keep_links=False
    )
    extracted_rows = 0
    try:
        sheet = workbook[config.workbook.canonical_sheet]
        for source_row, values in enumerate(
            sheet.iter_rows(
                min_row=config.sheet.first_data_row,
                max_col=len(SOURCE_HEADERS),
                values_only=True,
            ),
            start=config.sheet.first_data_row,
        ):
            if not any(value not in (None, "") for value in values):
                continue
            extracted_rows += 1
            yield RawDiagnosisRow.from_values(source_row, values)
        if extracted_rows != inspection.data_rows:
            raise WorkbookContractError(
                f"extracted {extracted_rows} rows after inspection reported {inspection.data_rows}"
            )
    finally:
        workbook.close()


def normalize_raw_diagnosis_row(
    raw: RawDiagnosisRow,
    source_version: str,
    source_sha256: str,
) -> DiagnosisRecord:
    main_code = _optional_text(raw.main_code, "main_code")
    additional_code = _optional_text(raw.additional_code, "additional_code")
    code = main_code or additional_code
    if code is None:
        raise DiagnosisNormalizationError("main_code or additional_code is required")
    return DiagnosisRecord(
        code=code,
        main_code=main_code,
        additional_code=additional_code,
        name=_required_text(raw.name, "name"),
        source_row=raw.source_row,
        source_version=_required_text(source_version, "source_version"),
        source_sha256=_required_text(source_sha256, "source_sha256"),
    )


def _required_text(value: RawCell, field_name: str) -> str:
    normalized = _optional_text(value, field_name)
    if normalized is None:
        raise DiagnosisNormalizationError(f"{field_name} is required")
    return normalized


def _optional_text(value: RawCell, field_name: str) -> str | None:
    if value is None:
        return None
    if not isinstance(value, str):
        raise DiagnosisNormalizationError(
            f"{field_name} must be text, found {type(value).__name__}"
        )
    normalized = unicodedata.normalize("NFC", value.strip())
    return normalized or None
