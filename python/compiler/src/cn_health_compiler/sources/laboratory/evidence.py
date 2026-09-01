"""Validate the pinned panel-evidence workbook without deriving clinical crosswalks."""

from pathlib import Path

from openpyxl import load_workbook


def inspect_panel_evidence(path: Path, expected_sheet: str) -> dict[int, str]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except (OSError, ValueError) as error:
        raise ValueError("panel evidence workbook is unreadable") from error
    try:
        if workbook.sheetnames != [expected_sheet]:
            raise ValueError("panel evidence worksheet identity changed")
        sheet = workbook[expected_sheet]
        if sheet.max_row != 667 or sheet.max_column != 53:
            raise ValueError("panel evidence workbook dimensions changed")
        row4 = tuple(sheet.cell(4, column).value for column in range(1, 10))
        row5 = tuple(sheet.cell(5, column).value for column in range(1, 11))
        if row4[:7] != (
            "序号",
            "项目名称",
            "加收项",
            "扩展项",
            "计价单位",
            "计价说明",
            "医保医疗服务项目分类与代码",
        ):
            raise ValueError("panel evidence primary headers changed")
        if row4[8] != "国家卫健委2023技术规范" or row5[6:10] != (
            "项目编码",
            "项目名称",
            "项目编码",
            "项目名称",
        ):
            raise ValueError("panel evidence mapping headers changed")
        names: dict[int, str] = {}
        for row_number in range(6, 668):
            ordinal = sheet.cell(row_number, 1).value
            name = sheet.cell(row_number, 2).value
            if ordinal != row_number - 5 or not isinstance(name, str) or not name.strip():
                raise ValueError(f"panel evidence data row changed: {row_number}")
            names[row_number] = name.strip()
        return names
    finally:
        workbook.close()
