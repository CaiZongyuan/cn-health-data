from hashlib import sha256
from pathlib import Path
from zipfile import ZipFile

import pytest
from cn_health_compiler.core.source import snapshot_local_source
from cn_health_compiler.sources.nhsa_drugs.workbook import (
    NhsaDrugWorkbookConfig,
    WorkbookContractError,
    inspect_workbook,
)
from openpyxl import Workbook

SHEETS = ("西药中成药新增变更", "本省双通道", "总表")
HEADERS = ("药品代码", "注册名称")


def _write_workbook(path: Path, *, formula: bool = False) -> None:
    workbook = Workbook()
    workbook.active.title = SHEETS[0]
    workbook.create_sheet(SHEETS[1])
    total = workbook.create_sheet(SHEETS[2])
    total.append(HEADERS)
    total.append(("XA01", "测试药品"))
    total.append(("XA02", '=CONCAT("测试", "公式")' if formula else "另一药品"))
    workbook.save(path)


def _config(path: Path, *, headers: tuple[str, ...] = HEADERS) -> NhsaDrugWorkbookConfig:
    with ZipFile(path) as archive:
        entry_count = len(archive.infolist())
        uncompressed_size = sum(item.file_size for item in archive.infolist())
    return NhsaDrugWorkbookConfig.model_validate(
        {
            "version": 1,
            "source": {
                "filename": path.name,
                "sha256": sha256(path.read_bytes()).hexdigest(),
                "size_bytes": path.stat().st_size,
            },
            "workbook": {
                "required_sheets": SHEETS,
                "canonical_sheet": "总表",
                "resolve_external_links": False,
            },
            "container": {
                "expected_zip_entries": entry_count,
                "expected_uncompressed_size_bytes": uncompressed_size,
                "max_uncompressed_size_bytes": uncompressed_size,
                "reject_macros": True,
            },
            "sheet": {
                "dimension": "A1:B3",
                "header_row": 1,
                "first_data_row": 2,
                "expected_data_rows": 2,
                "expected_formula_cells": 0,
                "headers": headers,
            },
        }
    )


def test_inspect_workbook_matches_declared_fingerprint(tmp_path: Path) -> None:
    source = tmp_path / "drugs.xlsx"
    _write_workbook(source)
    config = _config(source)
    snapshot = snapshot_local_source(source, config.source.sha256, tmp_path / "snapshots")

    inspection = inspect_workbook(snapshot, config)

    assert inspection.sheet_names == SHEETS
    assert inspection.dimension == "A1:B3"
    assert inspection.headers == HEADERS
    assert inspection.data_rows == 2
    assert inspection.formula_cells == 0


def test_inspect_workbook_rejects_formula_in_canonical_sheet(tmp_path: Path) -> None:
    source = tmp_path / "drugs.xlsx"
    _write_workbook(source, formula=True)
    config = _config(source)
    snapshot = snapshot_local_source(source, config.source.sha256, tmp_path / "snapshots")

    with pytest.raises(WorkbookContractError, match="formula cells"):
        inspect_workbook(snapshot, config)


def test_inspect_workbook_rejects_header_change(tmp_path: Path) -> None:
    source = tmp_path / "drugs.xlsx"
    _write_workbook(source)
    config = _config(source, headers=("药品代码", "错误名称"))
    snapshot = snapshot_local_source(source, config.source.sha256, tmp_path / "snapshots")

    with pytest.raises(WorkbookContractError, match="headers changed"):
        inspect_workbook(snapshot, config)
